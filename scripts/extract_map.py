#!/usr/bin/env python3
"""Convert the Store Map Google Sheet (xlsx export) into slots.json geometry.

Cells/merged ranges with values become slots (id like "61"/"86a") or labels.
Medium/thick cell borders become wall segments. Column widths / row heights
give true proportions.
"""
import json, re, sys
import openpyxl

XLSX = 'data/storemap.xlsx'
OUT = 'data/slots.json'
FLOORS = ['First floor', 'Second floor']
SLOT_RE = re.compile(r'^\d+[a-zA-Z]?$')
COL_PX = 7.0   # px per excel width unit
ROW_PX = 4/3   # px per point


def sheet_geometry(ws):
    ncols, nrows = ws.max_column, ws.max_row
    dcw = ws.sheet_format.defaultColWidth or 8.43
    drh = ws.sheet_format.defaultRowHeight or 15.0
    xs, x = [0.0], 0.0
    for c in range(1, ncols + 2):
        letter = openpyxl.utils.get_column_letter(c)
        cd = ws.column_dimensions.get(letter)
        w = cd.width if (cd and cd.width) else dcw
        x += w * COL_PX + 5
        xs.append(x)
    ys, y = [0.0], 0.0
    for r in range(1, nrows + 2):
        rd = ws.row_dimensions.get(r)
        h = rd.height if (rd and rd.height) else drh
        y += h * ROW_PX
        ys.append(y)
    return xs, ys


def fill_hex(cell):
    f = cell.fill
    if f and f.fill_type == 'solid' and f.fgColor and f.fgColor.type == 'rgb':
        rgb = f.fgColor.rgb
        if rgb and rgb not in ('00000000', 'FFFFFFFF'):
            return '#' + rgb[-6:].lower()
    return None


WALL_STYLES = {'medium', 'thick', 'double'}

def extract(ws):
    xs, ys = sheet_geometry(ws)
    merged = {}
    covered = set()
    for rng in ws.merged_cells.ranges:
        merged[(rng.min_row, rng.min_col)] = rng
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                covered.add((r, c))
    slots, labels = [], []
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if v is None or str(v).strip() == '':
                continue
            r, c = cell.row, cell.column
            if (r, c) in covered and (r, c) not in merged:
                continue  # non-anchor of merge
            if (r, c) in merged:
                rng = merged[(r, c)]
                x, y = xs[rng.min_col - 1], ys[rng.min_row - 1]
                w, h = xs[rng.max_col] - x, ys[rng.max_row] - y
            else:
                x, y = xs[c - 1], ys[r - 1]
                w, h = xs[c] - x, ys[r] - y
            if isinstance(v, float) and v == int(v):
                v = int(v)
            text = str(v).strip()
            item = {'x': round(x, 1), 'y': round(y, 1), 'w': round(w, 1), 'h': round(h, 1)}
            fill = fill_hex(cell)
            if fill:
                item['fill'] = fill
            if SLOT_RE.match(text):
                slots.append({'id': text, **item})
            else:
                labels.append({'text': text, **item})
    # wall segments from medium/thick borders
    walls = []
    for row in ws.iter_rows():
        for cell in row:
            r, c = cell.row, cell.column
            b = cell.border
            x0, x1 = xs[c - 1], xs[c]
            y0, y1 = ys[r - 1], ys[r]
            if b.top and b.top.style in WALL_STYLES:
                walls.append((x0, y0, x1, y0))
            if b.bottom and b.bottom.style in WALL_STYLES:
                walls.append((x0, y1, x1, y1))
            if b.left and b.left.style in WALL_STYLES:
                walls.append((x0, y0, x0, y1))
            if b.right and b.right.style in WALL_STYLES:
                walls.append((x1, y0, x1, y1))
    # merge collinear consecutive segments
    walls = sorted(set(walls))
    merged_walls, used = [], [False] * len(walls)
    hs = {}
    for seg in walls:
        x0, y0, x1, y1 = seg
        if y0 == y1:
            hs.setdefault(('h', y0), []).append((x0, x1))
        else:
            hs.setdefault(('v', x0), []).append((y0, y1))
    segs = []
    for (kind, k), spans in hs.items():
        spans.sort()
        cur0, cur1 = spans[0]
        for a, b2 in spans[1:]:
            if a <= cur1 + 0.5:
                cur1 = max(cur1, b2)
            else:
                segs.append((kind, k, cur0, cur1))
                cur0, cur1 = a, b2
        segs.append((kind, k, cur0, cur1))
    for kind, k, a, b2 in segs:
        if kind == 'h':
            merged_walls.append({'x1': round(a, 1), 'y1': round(k, 1), 'x2': round(b2, 1), 'y2': round(k, 1)})
        else:
            merged_walls.append({'x1': round(k, 1), 'y1': round(a, 1), 'x2': round(k, 1), 'y2': round(b2, 1)})
    # crop to content bbox
    items = slots + labels
    pts = [(s['x'], s['y']) for s in items] + [(s['x'] + s['w'], s['y'] + s['h']) for s in items] \
        + [(w['x1'], w['y1']) for w in merged_walls] + [(w['x2'], w['y2']) for w in merged_walls]
    minx = min(p[0] for p in pts) - 20; miny = min(p[1] for p in pts) - 20
    maxx = max(p[0] for p in pts) + 20; maxy = max(p[1] for p in pts) + 20
    for s in items:
        s['x'] = round(s['x'] - minx, 1); s['y'] = round(s['y'] - miny, 1)
    for w in merged_walls:
        w['x1'] = round(w['x1'] - minx, 1); w['x2'] = round(w['x2'] - minx, 1)
        w['y1'] = round(w['y1'] - miny, 1); w['y2'] = round(w['y2'] - miny, 1)
    return {'width': round(maxx - minx, 1), 'height': round(maxy - miny, 1),
            'slots': slots, 'labels': labels, 'walls': merged_walls}


def main():
    wb = openpyxl.load_workbook(XLSX)
    out = {'floors': []}
    for name in FLOORS:
        data = extract(wb[name])
        data['name'] = name
        out['floors'].append(data)
        print(f"{name}: {len(data['slots'])} slots, {len(data['labels'])} labels, "
              f"{len(data['walls'])} wall segments, {data['width']:.0f}x{data['height']:.0f}px")
    ids = [s['id'] for f in out['floors'] for s in f['slots']]
    dupes = {i for i in ids if ids.count(i) > 1}
    if dupes:
        print('NOTE: slot ids on both floors or duplicated in sheet:', sorted(dupes)[:20])
    with open(OUT, 'w') as fh:
        json.dump(out, fh)
    print('wrote', OUT)


if __name__ == '__main__':
    main()
